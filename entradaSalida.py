import openpyxl as op
import os.path

from pydantic   import BaseModel
from modelo     import Instructor, Ficha

# esta es una clase de la capa de infraestructua, lee y escribe datos en un archivo de excel
class EntradaSalida:
    def __init__(self):
        pass

    # retorna un diccionario {encabezado : numero} con las primeras filas de la hoja
    def getHeaders(self, archivo: op.Workbook, hoja: str) -> dict[str : int]:
        hoja = archivo[hoja]
        primeraFila = hoja.iter_rows(min_row=0,  max_row=1, values_only=True)
        encabezados = [b for b in  [a for a in primeraFila][0]]
        numeros = range(len(encabezados)) 

        return  dict(zip(encabezados, numeros))

    # lee los datos una hoja de excel y retorna una lista de objetos segun el modelo(BaseModel)
    def getData(self, nameArchivo: str, nameHoja: str, modelo = BaseModel):
        try:
            datos =[]
            archivo = op.load_workbook(os.path.join('datos', nameArchivo))
            headers  = self.getHeaders(archivo, nameHoja)

            hoja = archivo[nameHoja]
            fields_set = modelo.model_fields_set

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                dic = {}
                for header in headers:
                    dic[header.lower()] = fila[headers[header]]
                datos.append(modelo.model_construct(_fields_set=fields_set, **dic))

            return(datos)

        except KeyError as e:
            print('La hoja no existe', e )
        except FileNotFoundError as e:
            print('Archivo no existe ', e)
        finally:
            archivo.close
    
    # escribir una hoja en un archivo
    def writeSheet(self, nameArchivo: str, nameHoja: str, encabezados: list, filas: list):
        try:
            archivo = op.load_workbook(os.path.join('datos', nameArchivo))

            try:
               hojaSalida = archivo[nameHoja]

            except:
                hojaSalida = archivo.create_sheet(title=nameHoja)
                hojaSalida.append(encabezados)

            for fila in filas:
                hojaSalida.append(fila)

            archivo.save(os.path.join('datos', nameArchivo))

        except KeyError as e:
            print('La hoja no existe', e )
        except FileNotFoundError as e:
            print('Archivo no existe ', e)
        finally:
            archivo.close

if __name__ == "__main__":
    # instructores = EntradaSalida().getData('consolidado.xlsx', 'instructores', Instructor)
    # for fila in instructores:
    #     print(fila)
    fichas = EntradaSalida().getData('consolidado.xlsx', 'fichasD', Ficha)
    for fila in fichas:
        print(fila)
    # EntradaSalida().writeSheet('consolidado.xlsx', 'salida', ['cedula','nombre','competencia'], [[890,"Ins1","MAT"],[891,"Ins2","ETI"]])